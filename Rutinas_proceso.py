import openpyxl
import os

from PyQt5 import QtWidgets
from openpyxl.utils.exceptions import InvalidFileException


class BidClient:

    def __init__(self, client='', bids='', id_contract='', descr_proj='', descr_jira='', am_email='', fin_contrato='',
                 due_date='', fecha_apertura='', acv='', sa_email=''):
        self.client = client
        self.bids = bids


def obtener_clientes_agrupados(hoja):
    ultima_fila = hoja.max_row - 1
    columna_clientes = hoja['B2': 'B' + str(ultima_fila)]  # Extraemos la columna de los clientes
    lista_general_clientes = [item[0].value
                              for item in columna_clientes]

    # El indice_separación marca la frontera entre líneas agrupables y no agrupables
    indice_separacion = lista_general_clientes.index('No unificar en adelante')
    lista_clientes_agrupados = lista_general_clientes[:indice_separacion]

    lista_bids = [item[0].value
                  for item in hoja['C2': 'C' + str(indice_separacion + 1)]]  # Lista de BIDs
    lista_id_contract = [item[0].value
                         for item in hoja['A2': 'A' + str(indice_separacion + 1)]]  # Lista de números de contrato
    lista_descr_proj = [item[0].value
                        for item in hoja['E2': 'E' + str(indice_separacion + 1)]]  # Lista de descripciones de proyecto
    lista_descr_jira = [item[0].value
                        for item in hoja['F2': 'F' + str(indice_separacion + 1)]]  # Lista de descripciones de JIRA
    lista_am_email = [item[0].value
                      for item in hoja['J2': 'J' + str(indice_separacion + 1)]]  # Lista de e-mails de AM
    lista_fin_contrato = [item[0].value
                          for item in hoja['K2': 'K' + str(indice_separacion + 1)]]  # Lista de finales de contrato
    lista_due_date = [item[0].value
                      for item in hoja['N2': 'N' + str(indice_separacion + 1)]]  # Lista de descripciones de due date
    lista_fecha_apertura = [item[0].value
                            for item in hoja['O2': 'O' + str(indice_separacion + 1)]]  # Lista de fechas de apertura
    lista_ACV = [item[0].value
                 for item in hoja['P2': 'P' + str(indice_separacion + 1)]]  # Lista de ACV
    lista_email_SA_asignado = [item[0].value
                               for item in hoja['S2': 'S' + str(indice_separacion + 1)]]  # Lista de SA asignado

    set_clientes = sorted(set(lista_clientes_agrupados))  # Quedan los nombres de clientes sin repetición
    lista_clientes_y_bids = []

    for x in set_clientes:  # Creamos una lista de clases BidClient
        customer = BidClient(x)
        lista_clientes_y_bids.append(customer)

    for item in lista_clientes_y_bids:

        # Ahora definimos una serie de objetos auxiliares que nos servirán para elegir los datos usados en
        # la agrupación. El campo fundamental es
        string_bids = ''
        aux_bids = []
        aux_contract_id = []
        aux_descr_proj = []
        aux_descr_jira = []
        aux_am_email = []
        aux_fin_contr = []
        aux_due_date = []
        aux_fecha_apertura = []
        aux_acv = []
        aux_sa_asign = []

        for i in range(len(lista_clientes_agrupados)):
            if item.client == lista_clientes_agrupados[i]:
                aux_bids.append(lista_bids[i])
                aux_contract_id.append(lista_id_contract[i])
                aux_descr_proj.append(lista_descr_proj[i])
                aux_descr_jira.append(lista_descr_jira[i])
                aux_am_email.append(lista_am_email[i])
                aux_fin_contr.append(lista_fin_contrato[i].date())
                aux_due_date.append(lista_due_date[i].date())
                aux_fecha_apertura.append(lista_fecha_apertura[i].date())
                aux_acv.append(lista_ACV[i])
                aux_sa_asign.append(lista_email_SA_asignado[i])

        # Ahora tratamos la lista de bids para transformarla en un set y luego en un string
        set_aux_bids = sorted(set(aux_bids))

        for bid in set_aux_bids:
            string_bids += (bid + ', ')

        string_bids = string_bids[0:len(string_bids) - 2]  # Esto es para quitar la coma y el blanco del último ítem
        item.bids = string_bids
        # Ahora calculamos el ítem con la fecha más próxima, porque el resto de campos van a tomarse de él

        idx_ok = aux_fin_contr.index(min(aux_fin_contr))
        item.id_contract = aux_contract_id[idx_ok]
        item.descr_proj = aux_descr_proj[idx_ok]
        item.descr_jira = aux_descr_jira[idx_ok]
        item.am_email = aux_am_email[idx_ok]
        item.fin_contrato = aux_fin_contr[idx_ok]
        item.due_date = aux_due_date[idx_ok]
        item.fecha_apertura = aux_fecha_apertura[idx_ok]
        suma_acv = 0.0
        for acv_item in aux_acv:
            suma_acv += acv_item
        item.acv = suma_acv
        item.sa_email = aux_sa_asign[idx_ok]

    return lista_clientes_y_bids


def obtener_clientes_no_agrupados(hoja):
    ultima_fila = hoja.max_row - 1

    lista_general_clientes = [item[0].value
                              for item in hoja['B2':'B' + str(ultima_fila)]]  # Lista completa de clientes

    # El indice_separación marca la frontera entre líneas agrupables y no agrupables
    indice_separacion = lista_general_clientes.index('No unificar en adelante')

    lista_clientes_no_agrupados = lista_general_clientes[indice_separacion + 1:ultima_fila + 1]
    lista_bids = [item[0].value
                  for item in hoja['C' + str(indice_separacion + 3): 'C' + str(ultima_fila)]]  # Lista de BIDs
    lista_id_contract = [item[0].value
                         for item in
                         hoja['A' + str(indice_separacion + 3):'A' + str(ultima_fila)]]  # Lista de números de contrato
    lista_descr_proj = [item[0].value
                        for item in hoja['E' + str(indice_separacion + 3):'E' + str(
            ultima_fila)]]  # Lista de descripciones de proyecto
    lista_descr_jira = [item[0].value
                        for item in
                        hoja['F' + str(indice_separacion + 3):'F' + str(ultima_fila)]]  # Lista de descripciones de JIRA
    lista_am_email = [item[0].value
                      for item in
                      hoja['J' + str(indice_separacion + 3):'J' + str(ultima_fila)]]  # Lista de e-mails de AM
    lista_fin_contrato = [item[0].value
                          for item in
                          hoja['K' + str(indice_separacion + 3):'K' + str(ultima_fila)]]  # Lista de finales de contrato
    lista_due_date = [item[0].value
                      for item in hoja['N' + str(indice_separacion + 3):'N' + str(
            ultima_fila)]]  # Lista de descripciones de due date
    lista_fecha_apertura = [item[0].value
                            for item in hoja['O' + str(indice_separacion + 3):'O' + str(
            ultima_fila)]]  # Lista de fechas de apertura
    lista_ACV = [item[0].value
                 for item in hoja['P' + str(indice_separacion + 3):'P' + str(ultima_fila)]]  # Lista de ACV
    lista_email_SA_asignado = [item[0].value
                               for item in
                               hoja['R' + str(indice_separacion + 3):'R' + str(ultima_fila)]]  # Lista de SA asignado

    set_bids = set(lista_bids)  # Obtenemos un listado de BIDs no repetidas

    lista_bids_y_clientes = []

    for x in set_bids:  # Creamos una lista de clases BidClient
        bid = BidClient('Ful', x)
        lista_bids_y_clientes.append(bid)

    for item in lista_bids_y_clientes:  # Vamos procesando por BID en vez de por cliente como el la lista de agrupados

        # Ahora definimos una serie de objetos auxiliares que nos servirán para elegir los datos usados en
        # la agrupación. El campo fundamental es de la BID
        aux_customer = []
        aux_contract_id = []
        aux_descr_proj = []
        aux_descr_jira = []
        aux_am_email = []
        aux_fin_contr = []
        aux_due_date = []
        aux_fecha_apertura = []
        aux_acv = []
        aux_sa_asign = []

        for i in range(len(lista_bids)):
            if item.bids == lista_bids[i]:  # Vamos adquiriendo los datos de las líneas con el BID en cuestión
                aux_contract_id.append(lista_id_contract[i])
                aux_customer.append(lista_clientes_no_agrupados[i])
                aux_descr_proj.append(lista_descr_proj[i])
                aux_descr_jira.append(lista_descr_jira[i])
                aux_am_email.append(lista_am_email[i])
                aux_fin_contr.append(lista_fin_contrato[i].date())
                aux_due_date.append(lista_due_date[i].date())
                aux_fecha_apertura.append(lista_fecha_apertura[i].date())
                aux_acv.append(lista_ACV[i])
                aux_sa_asign.append(lista_email_SA_asignado[i])

        idx_ok = aux_fin_contr.index(min(aux_fin_contr))
        item.id_contract = aux_contract_id[idx_ok]
        item.client = aux_customer[0]
        item.descr_proj = aux_descr_proj[idx_ok]
        item.descr_jira = aux_descr_jira[idx_ok]
        item.am_email = aux_am_email[idx_ok]
        item.fin_contrato = aux_fin_contr[idx_ok]
        item.due_date = aux_due_date[idx_ok]
        item.fecha_apertura = aux_fecha_apertura[idx_ok]
        suma_acv = 0.0
        for acv_item in aux_acv:
            suma_acv += acv_item
        item.acv = suma_acv
        item.sa_email = aux_sa_asign[idx_ok]

    lista_bids_y_clientes = sorted(lista_bids_y_clientes, key=lambda bid: (bid.client, bid.bids))

    return lista_bids_y_clientes


def escribir_clientes_en_Excel(lista_clientes, hoja_destino, first_line):
    line = first_line

    for cliente in lista_clientes:
        hoja_destino['A' + str(line)] = cliente.id_contract
        hoja_destino['B' + str(line)] = cliente.client
        hoja_destino['C' + str(line)] = cliente.bids
        hoja_destino['D' + str(line)] = cliente.descr_proj
        hoja_destino['E' + str(line)] = cliente.descr_jira
        hoja_destino['F' + str(line)] = cliente.am_email
        hoja_destino['G' + str(line)] = cliente.fin_contrato
        hoja_destino['H' + str(line)] = cliente.due_date
        hoja_destino['I' + str(line)] = cliente.fecha_apertura
        hoja_destino['J' + str(line)] = cliente.acv
        hoja_destino['K' + str(line)] = cliente.sa_email

        line += 1


def procesar_fichero(hoja_origen, hoja_destino, padre):
    """
    Este procedimiento hace el trabajo de procesado a partir de dos hojas Excel, la de origen y la de resultado
    final (vacía a la entrada). Para ello invoca otros procedimientos subordinados
    :param hoja_origen: Hoja Excel ya rellena por el usuario en el fichero de entrada
    :param hoja_destino: Hoja Excel vacía a la entrada y en la que se depositará el resultado final
    :param padre: Instancia de la pantalla principal
    :return: hoja_destino rellena
    """
    padre.signals.informacion.emit('Obteniendo lista de clientes agrupados')
    lista_clientes_agrupados = obtener_clientes_agrupados(hoja_origen)
    padre.signals.informacion.emit('Escribiendo clientes agrupados en Excel')
    escribir_clientes_en_Excel(lista_clientes_agrupados, hoja_destino, 2)
    padre.signals.informacion.emit('Obteniendo lista de clientes no agrupados')
    lista_clientes_no_agrupados = obtener_clientes_no_agrupados(hoja_origen)
    first_line = 4 + len(lista_clientes_agrupados)
    padre.signals.informacion.emit('Escribiendo clientes no agrupados en Excel')

    print('Escribiendo clientes no agrupados en Excel')
    escribir_clientes_en_Excel(lista_clientes_no_agrupados, hoja_destino, first_line)
    padre.signals.informacion.emit('Salvando fichero de renovaciones')


def procesamiento_general(nombre_fichero, padre):
    """
    Esta es la rutina principal de la aplicación. Convierte el fichero Excel de entrada en otro, llamado Renovaciones
    en el que se han consolidado los datos según especificaciones
    :param nombre_fichero: Nombre completo del fichero Excel a procesar
    :param padre: Instancia de la clase principal
    :return: Fichero Renovaciones.xlsx
    """

    try:
        print('Abriendo libro')
        padre.signals.informacion.emit('Abriendo libro')
        libro = openpyxl.load_workbook(nombre_fichero, data_only=True)
        print('abriendo hoja')
        hoja_origen = libro.active
        hoja_destino = libro['Salida JIRA']

    except InvalidFileException:
        print('No Excel')
        padre.signals.error.emit('No es un fichero Excel')
    except KeyError:
        padre.signals.error.emit('No existe la página "Salida JIRA"')

    else:  # LLegados aquí ya sabemos que no se ha producido una excepción. Procesamos el fichero
        procesar_fichero(hoja_origen, hoja_destino, padre)
        carpeta, nombre_file = os.path.split(nombre_fichero)

        file_out = os.path.join(carpeta, 'Renovaciones.xlsx')

        terminado = False
        while not terminado:
            try:
                libro.save(file_out)
            except PermissionError:
                QtWidgets.QMessageBox.critical(padre, "Error", "Fichero Renovaciones abierto \n Ciérrelo para seguir")
            else:
                terminado = True

        padre.signals.completar_libro.emit()




